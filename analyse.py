import json
import os
from typing import List, Dict
from datetime import datetime
import re
import pandas as pd
import requests
from io import BytesIO
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Side, Border, Font
from PIL import Image as PILImage  # 确保 Pillow 已安装
import cv2
import numpy as np
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import base64
import time
from bs4 import BeautifulSoup
import hashlib
from urllib.parse import urlencode

class ImageMatcher:
    def __init__(self):
        self.driver = None
        self.similarity_threshold = 0.8  # 相似度阈值
        self.alibaba_searcher = Alibaba1688Searcher()

    def init_browser(self):
        """初始化浏览器"""
        self.alibaba_searcher.init_browser()

    def close_browser(self):
        """关闭浏览器"""
        self.alibaba_searcher.close_browser()

    def search_1688_by_image(self, image_url: str) -> List[Dict]:
        """
        使用图片在1688上搜索
        返回搜索结果列表，每个结果包含图片URL和商品链接
        """
        return self.alibaba_searcher.search_by_image(image_url)

    def compare_images(self, img1_url: str, img2_url: str) -> float:
        """
        比较两张图片的相似度
        返回相似度得分(0-1)
        """
        # 下载图片
        img1 = self._download_and_process_image(img1_url)
        img2 = self._download_and_process_image(img2_url)
        
        # 使用OpenCV计算图片相似度
        try:
            # 转换为相同大小
            img1 = cv2.resize(img1, (224, 224))
            img2 = cv2.resize(img2, (224, 224))
            
            # 计算直方图
            hist1 = cv2.calcHist([img1], [0, 1, 2], None, [8, 8, 8], [0, 256, 0, 256, 0, 256])
            hist2 = cv2.calcHist([img2], [0, 1, 2], None, [8, 8, 8], [0, 256, 0, 256, 0, 256])
            
            # 归一化直方图
            cv2.normalize(hist1, hist1)
            cv2.normalize(hist2, hist2)
            
            # 计算相似度
            similarity = cv2.compareHist(hist1, hist2, cv2.HISTCMP_CORREL)
            return similarity
        except Exception as e:
            print(f"比较图片时出错: {str(e)}")
            return 0.0

    def _download_and_process_image(self, url: str) -> np.ndarray:
        """下载并处理图片"""
        response = requests.get(url)
        img_array = np.asarray(bytearray(response.content), dtype=np.uint8)
        return cv2.imdecode(img_array, cv2.IMREAD_COLOR)

class Alibaba1688Searcher:
    def __init__(self):
        self.session = requests.Session()
        self.base_url = "https://re.1688.com"
        self.api_url = "https://h5api.m.1688.com/h5"
        self.driver = None
        self.m_h5_tk = None  # 存储token
        self.headers = None  # 存储请求头

    def init_browser(self):
        """初始化浏览器和必要的参数"""
        # 初始化浏览器
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')
        options.add_argument('--disable-gpu')
        options.add_argument('--no-sandbox')
        self.driver = webdriver.Chrome(options=options)
        
        # 初始化token和请求头
        self._init_token_and_headers()

    def _init_token_and_headers(self):
        """初始化token和请求头（只需执行一次）"""
        try:
            print("\n1. 初始化token和请求头...")
            
            # 访问1688识图页面
            self.driver.get(self.base_url)
            
            # 等待页面加载完成
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.TAG_NAME, "body"))
            )
            
            # 获取cookies和localStorage
            cookies = self.driver.get_cookies()
            token = self.driver.execute_script('return localStorage.getItem("_m_h5_tk");')
            
            # 设置cookies
            for cookie in cookies:
                self.session.cookies.set(cookie['name'], cookie['value'])
                if cookie['name'] == '_m_h5_tk':
                    self.m_h5_tk = cookie['value'].split('_')[0]
            
            # 设置通用请求头
            self.headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36',
                'Accept': 'application/json',
                'Accept-Language': 'zh-CN,zh;q=0.9',
                'Content-Type': 'application/x-www-form-urlencoded',
                'Origin': 'https://re.1688.com',
                'Referer': 'https://re.1688.com/',
                'Cache-Control': 'no-cache',
                'Pragma': 'no-cache',
                'sec-ch-ua': '"Chromium";v="134", "Not:A-Brand";v="24", "Google Chrome";v="134"',
                'sec-ch-ua-mobile': '?0',
                'sec-ch-ua-platform': '"Windows"',
                'sec-fetch-dest': 'empty',
                'sec-fetch-mode': 'cors',
                'sec-fetch-site': 'same-site'
            }
            
            print("✓ token和请求头初始化成功")
            
        except Exception as e:
            print(f"× 初始化token和请求头失败: {str(e)}")
            raise

    def close_browser(self):
        """关闭浏览器"""
        if self.driver:
            self.driver.quit()

    def _image_to_base64(self, image_url: str) -> str:
        """将图片转换为base64编码，增强错误处理和重试机制"""
        max_retries = 3
        
        for attempt in range(max_retries):
            try:
                print(f"尝试下载图片 (第 {attempt+1}/{max_retries} 次)...")
                print(f"图片URL: {image_url}")
                
                # 构建更完整的请求头
                headers = {
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
                    'Accept': 'image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8',
                    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
                    'Referer': 'https://echotik.live/',  # 更新为正确的referer
                    'Origin': 'https://echotik.live',
                    'sec-ch-ua': '"Chromium";v="134", "Not:A-Brand";v="24", "Google Chrome";v="134"',
                    'sec-ch-ua-mobile': '?0',
                    'sec-ch-ua-platform': '"Windows"',
                    'sec-fetch-dest': 'image',
                    'sec-fetch-mode': 'no-cors',
                    'sec-fetch-site': 'cross-site'
                }
                
                # 尝试直接下载
                response = requests.get(image_url, headers=headers, timeout=15)
                
                # 检查响应状态
                if response.status_code == 200:
                    print(f"图片下载成功，内容类型: {response.headers.get('Content-Type', '未知')}")
                    print(f"图片大小: {len(response.content)} 字节")
                    
                    # 将图片内容转换为base64
                    base64_data = base64.b64encode(response.content).decode('utf-8')
                    print(f"Base64数据长度: {len(base64_data)}")
                    return base64_data
                else:
                    print(f"下载图片失败: HTTP {response.status_code}")
                    
                    # 如果是CDN问题，尝试替代方案
                    if response.status_code == 403 and attempt < max_retries - 1:
                        print("尝试替代下载方法...")
                        
                        # 方法1: 尝试使用不同的referer
                        alt_headers = headers.copy()
                        alt_headers['Referer'] = image_url
                        
                        try:
                            alt_response = requests.get(image_url, headers=alt_headers, timeout=15)
                            if alt_response.status_code == 200:
                                print("替代方法1成功!")
                                base64_data = base64.b64encode(alt_response.content).decode('utf-8')
                                return base64_data
                        except:
                            pass
                        
                        # 方法2: 如果是echotik.shop的图片，尝试修改URL
                        if 'echotik.shop' in image_url:
                            try:
                                # 提取图片ID部分
                                parts = image_url.split('/')
                                if len(parts) >= 5:
                                    img_id = parts[-2] + '/' + parts[-1]
                                    alt_url = f"https://echotik.live/storage/products/{img_id}"
                                    print(f"尝试替代URL: {alt_url}")
                                    
                                    alt_response = requests.get(alt_url, headers=headers, timeout=15)
                                    if alt_response.status_code == 200:
                                        print("替代方法2成功!")
                                        base64_data = base64.b64encode(alt_response.content).decode('utf-8')
                                        return base64_data
                            except:
                                pass
                    
                    # 等待后重试
                    if attempt < max_retries - 1:
                        wait_time = 2 * (attempt + 1)  # 指数退避
                        print(f"等待 {wait_time} 秒后重试...")
                        time.sleep(wait_time)
                
            except Exception as e:
                print(f"下载图片时出错: {str(e)}")
                if attempt < max_retries - 1:
                    wait_time = 2 * (attempt + 1)
                    print(f"等待 {wait_time} 秒后重试...")
                    time.sleep(wait_time)
        
        # 所有尝试都失败，返回一个默认图片的base64
        print("所有下载尝试都失败，使用默认图片")
        try:
            # 使用一个简单的1x1像素透明PNG作为默认图片
            default_img = "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNkYAAAAAYAAjCB0C8AAAAASUVORK5CYII="
            return default_img
        except:
            return None

    def download_image(self, url: str) -> bytes:
        """下载图片"""
        max_retries = 3
        retry_delay = 2  # 初始重试延迟（秒）
        
        print(f"\n开始处理图片: {url}")
        
        for attempt in range(max_retries):
            try:
                print(f"尝试下载图片 (第 {attempt+1}/{max_retries} 次)...")
                print(f"图片URL: {url}")
                
                response = requests.get(url, timeout=10)
                if response.status_code == 200:
                    return response.content
                else:
                    print(f"下载图片失败: HTTP {response.status_code}")
                    
                    # 尝试替代URL
                    if 'cdn.echotik.shop' in url:
                        alt_url = url.replace('cdn.echotik.shop', 'echotik.live/storage')
                        print(f"尝试替代URL: {alt_url}")
                        alt_response = requests.get(alt_url, timeout=10)
                        if alt_response.status_code == 200:
                            return alt_response.content
                
                if attempt < max_retries - 1:
                    wait_time = retry_delay * (attempt + 1)
                    print(f"等待 {wait_time} 秒后重试...")
                    time.sleep(wait_time)
                
            except Exception as e:
                print(f"下载出错: {str(e)}")
                if attempt < max_retries - 1:
                    wait_time = retry_delay * (attempt + 1)
                    print(f"等待 {wait_time} 秒后重试...")
                    time.sleep(wait_time)
        
        # 所有尝试都失败
        raise Exception(f"无法下载图片: {url}")

    def upload_image(self, image_url: str) -> dict:
        """上传图片到1688"""
        try:
            # 1. 确保已初始化
            if not self.m_h5_tk or not self.headers:
                self._init_token_and_headers()
            
            # 2. 准备上传数据
            print(f"\n开始处理图片: {image_url}")
            timestamp = str(int(time.time() * 1000))
            image_base64 = self._image_to_base64(image_url)
            if not image_base64:
                raise Exception("图片转base64失败")

            data = {
                "imageBase64": image_base64,
                "appName": "searchImageUpload",
                "appKey": "pvvljh1grxcmaay2vgpe9nb68gg9ueg2"
            }
            
            # 3. 计算签名
            sign_content = f"{self.m_h5_tk}&{timestamp}&12574478&{json.dumps(data)}"
            sign = hashlib.md5(sign_content.encode('utf-8')).hexdigest()
            
            # 4. 构建请求参数
            params = {
                'jsv': '2.6.1',
                'appKey': '12574478',
                't': timestamp,
                'sign': sign,
                'api': 'mtop.1688.imageService.putImage',
                'v': '1.0',
                'type': 'originaljson',
                'dataType': 'jsonp',
                'timeout': '20000',
                'ignoreLogin': 'true',
                'prefix': 'h5api',
                'ecode': '0',
                'jsonpIncPrefix': 'search1688'
            }
            
            # 5. 发送请求
            upload_url = f"{self.api_url}/mtop.1688.imageservice.putimage/1.0/?{urlencode(params)}"
            
            response = self.session.post(
                upload_url,
                data={'data': json.dumps(data)},
                headers=self.headers
            )

            if response.status_code == 200:
                try:
                    result = response.json()
                    if result.get("ret", [""])[0].startswith("SUCCESS"):
                        return result
                    else:
                        raise Exception(f"上传失败: {result.get('ret', ['未知错误'])[0]}")
                except json.JSONDecodeError:
                    raise Exception("解析响应失败")
            else:
                raise Exception(f"请求失败: HTTP {response.status_code}")

        except Exception as e:
            print(f"\n上传图片时出错: {str(e)}")
            raise  # 向上传递错误

    def search_similar_products(self, image_id: str) -> List[Dict]:
        """搜索相似商品"""
        try:
            # 构建搜索URL
            search_url = f"https://s.1688.com/p4p/image/index.html?imageId={image_id}"
            self.driver.get(search_url)
            
            # 等待搜索结果加载
            WebDriverWait(self.driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, "sm-offer"))
            )
            
            # 解析搜索结果
            soup = BeautifulSoup(self.driver.page_source, 'html.parser')
            products = []
            
            for offer in soup.select(".normalcommon-offer-card"):
                try:
                    # 提取商品链接
                    link_elem = offer.select_one("a[href*='detail.1688.com']")
                    if not link_elem:
                        continue
                    product_url = link_elem.get('href', '')
                    
                    # 提取商品图片
                    img_div = offer.select_one(".img")
                    if not img_div:
                        continue
                    style = img_div.get('style', '')
                    image_url = ''
                    if 'background-image' in style:
                        image_url = style.split('url(')[1].split(')')[0].strip('"')
                    
                    # 提取商品标题
                    title_elem = offer.select_one(".title")
                    title = title_elem.text.strip() if title_elem else ''
                    
                    # 提取价格
                    price_elem = offer.select_one(".price")
                    price = price_elem.text.strip() if price_elem else ''
                    
                    products.append({
                        'product_url': product_url,
                        'image_url': image_url,
                        'title': title,
                        'price': price
                    })
                    
                    if len(products) >= 10:  # 只取前10个结果
                        break
                        
                except Exception as e:
                    print(f"解析商品数据时出错: {str(e)}")
                    continue
            
            return products
            
        except Exception as e:
            print(f"搜索相似商品时出错: {str(e)}")
            return []

    def search_by_image(self, image_url: str) -> List[Dict]:
        """使用图片在1688上搜索"""
        try:
            print("\n=== 开始图片搜索流程 ===")
            print(f"1. 图片URL: {image_url}")
            
            print("\n2. 准备上传图片...")
            upload_result = self.upload_image(image_url)
            print(f"上传结果: {json.dumps(upload_result, indent=2, ensure_ascii=False) if upload_result else 'None'}")
            
            if not upload_result:
                print("× 图片上传失败，终止搜索")
                raise Exception("图片上传失败")
            
            # 从data字段中获取imageId
            image_id = upload_result.get("data", {}).get("imageId")
            if not image_id:
                print("× 未获取到imageId，终止搜索")
                raise Exception("未获取到imageId")
            
            print(f"\n3. 获取到imageId: {image_id}")
            
            print("\n4. 构建搜索URL...")
            search_url = (
                f"https://s.1688.com/p4p/image/index.html?"
                f"tab=imageSearch&"
                f"imageAddress=&"
                f"cosite=re&"
                f"keywordid=&"
                f"trackid=&"
                f"location=&"
                f"ptid=&"
                f"bd_vid=&"
                f"imageId={image_id}&"
                f"spm=a2638t.b_30496503.submit.input&"
                f"imageIdList={image_id}"
            )
            print(f"搜索URL: {search_url}")
            
            print("\n5. 使用Selenium访问搜索页面...")
            self.driver.get(search_url)
            print(f"当前页面URL: {self.driver.current_url}")
            print(f"当前页面标题: {self.driver.title}")
            
            print("\n6. 等待页面加载...")
            max_retries = 5
            offers = []
            
            for attempt in range(max_retries):
                try:
                    print(f"\n尝试 {attempt+1}/{max_retries} 次加载商品数据")
                    time.sleep(3)
                    
                    print("- 滚动页面触发懒加载")
                    self.driver.execute_script("window.scrollTo(0, 1000);")
                    time.sleep(1)
                    self.driver.execute_script("window.scrollTo(0, 2000);")
                    time.sleep(1)
                    
                    print("- 查找商品元素")
                    offers = self.driver.find_elements(By.CLASS_NAME, "normalcommon-offer-card")
                    
                    if offers and len(offers) > 0:
                        print(f"✓ 成功找到 {len(offers)} 个商品元素")
                        break
                    else:
                        print("× 未找到商品元素，准备重试")
                
                except Exception as e:
                    print(f"× 加载失败: {str(e)}")
                    if attempt == max_retries - 1:
                        raise Exception("页面加载超时，未找到商品元素")
                    time.sleep(2)
            
            print("\n7. 开始解析商品数据...")
            products = []
            
            print(f"准备解析 {len(offers[:10])} 个商品")
            for i, offer in enumerate(offers[:10], 1):
                try:
                    print(f"\n解析第 {i} 个商品:")
                    
                    # 找到所有链接，选择第一个包含详情页URL的链接
                    links = offer.find_elements(By.TAG_NAME, "a")
                    product_url = None
                    for link in links:
                        href = link.get_attribute('href')
                        if href and ('detail.1688.com' in href or 'dj.1688.com' in href):
                            product_url = href
                            break
                    
                    if not product_url:
                        print("× 未找到商品链接")
                        continue
                    print(f"  链接: {product_url}")
                    
                    print("- 获取商品图片")
                    img_div = offer.find_element(By.CSS_SELECTOR, ".img-container .img")
                    style = img_div.get_attribute('style')
                    image_url = ''
                    if 'url(' in style:
                        match = re.search(r'url\(["\']?(.*?)["\']?\)', style)
                        if match:
                            image_url = match.group(1)
                    print(f"  图片: {image_url}")
                    
                    print("- 获取标题和价格")
                    title = offer.find_element(By.CSS_SELECTOR, ".mojar-element-title .title").text
                    price = offer.find_element(By.CSS_SELECTOR, ".mojar-element-price .price").text
                    print(f"  标题: {title}")
                    print(f"  价格: {price}")
                    
                    if product_url and image_url and title:
                        products.append({
                            'title': title,
                            'price': price,
                            'product_url': product_url,
                            'image_url': image_url
                        })
                        print("✓ 商品解析成功")
                    else:
                        print("× 商品数据不完整")
                    
                except Exception as e:
                    print(f"× 解析商品数据时出错: {str(e)}")
                    import traceback
                    print(traceback.format_exc())
                    continue
            
            print(f"\n=== 搜索完成，共解析 {len(products)} 个商品 ===")
            if len(products) == 0:
                print("警告：未成功解析任何商品！")
            return products
            
        except Exception as e:
            print(f"\n=== 搜索过程出错 ===")
            print(f"错误信息: {str(e)}")
            import traceback
            print("\n详细错误信息:")
            print(traceback.format_exc())
            return []

class DataAnalyzer:
    def __init__(self, data_dir: str = 'data'):
        """初始化数据分析器
        
        Args:
            data_dir: 数据文件目录，默认为 'data'
        """
        self.data_dir = data_dir
        self.image_dir = 'product_images'  # 添加图片目录
        os.makedirs(self.image_dir, exist_ok=True)
        self.image_matcher = ImageMatcher()
        self.alibaba_searcher = Alibaba1688Searcher()
        
    def _load_latest_data(self) -> List[Dict]:
        """加载最新的数据文件"""
        try:
            # 获取所有products开头的json文件
            data_files = [
                f for f in os.listdir(self.data_dir) 
                if f.startswith('products_') and f.endswith('.json')
            ]
            
            if not data_files:
                raise FileNotFoundError("没有找到数据文件")
            
            # 按文件名排序，获取最新的文件
            latest_file = max(data_files)
            file_path = os.path.join(self.data_dir, latest_file)
            
            print(f"加载数据文件: {latest_file}")
            
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                print(f"成功加载 {len(data)} 条商品数据")
                return data
            
        except Exception as e:
            print(f"加载数据文件时出错: {str(e)}")
            raise
    
    def _convert_price(self, price_str: str) -> float:
        """将价格字符串转换为数字"""
        if not price_str or price_str == 'N/A':
            return 0.0
        # 提取数字部分（去掉RM前缀）
        match = re.search(r'RM([\d.]+)', price_str)
        if match:
            return float(match.group(1))
        return 0.0
    
    def _convert_count(self, count_str: str) -> float:
        """将数量字符串（如 "1.5K", "2.3M"）转换为数字"""
        if not count_str or count_str == 'N/A':
            return 0.0
            
        multipliers = {'K': 1000, 'M': 1000000}
        match = re.search(r'([\d.]+)([KM])?', count_str)
        if match:
            number = float(match.group(1))
            unit = match.group(2)
            if unit and unit in multipliers:
                number *= multipliers[unit]
            return number
        return 0.0
    
    def _download_image(self, url: str, product_id: str) -> str:
        """下载商品图片并返回本地路径，添加超时处理"""
        try:
            response = requests.get(url, timeout=10)  # 添加10秒超时
            if response.status_code == 200:
                image_path = os.path.join(self.image_dir, f"{product_id}.jpg")
                with open(image_path, 'wb') as f:
                    f.write(response.content)
                return image_path
        except requests.exceptions.Timeout:
            print(f"\n下载超时: {url}")
        except Exception as e:
            print(f"\n下载失败 {url}: {str(e)}")
        return None
    
    def analyze_products(self, 
                        price_range: tuple = (20, 50),    # 价格范围
                        sales_weight: float = 0.6,        # 销量得分权重
                        influencer_weight: float = 0.4,   # 达人数量得分权重
                        top_n: int = 50                   # 返回前N个结果
                        ) -> List[Dict]:
        """
        分析商品数据，使用排名评分系统
        
        策略：
        1. 首先筛选价格范围内的有效商品
        2. 对7天销量进行排名评分（由高到低）
        3. 对达人数量进行排名评分（由低到高）
        4. 计算加权总分并排序
        """
        products = self._load_latest_data()
        valid_products = []
        skipped_count = 0
        
        # 第一步：筛选价格范围内的有效商品
        for product in products:
            try:
                price = self._convert_price(product.get('avg_price', 'N/A'))
                sales = self._convert_count(product.get('total_sale_nd_cnt', 'N/A'))
                influencers = self._convert_count(product.get('influencers_count', 'N/A'))
                
                # 只保留价格范围内且数据有效的商品
                if (price_range[0] <= price <= price_range[1] and 
                    sales > 0 and influencers > 0):
                    valid_products.append({
                        'product': product,
                        'price': price,
                        'sales': sales,
                        'influencers': influencers
                    })
                else:
                    skipped_count += 1
                
            except Exception as e:
                print(f"处理商品数据时出错: {str(e)}")
                skipped_count += 1
                continue
        
        if not valid_products:
            print("\n警告: 没有找到有效的商品数据！")
            return []
        
        total_products = len(valid_products)
        score_step = 100 / total_products  # 计算分数间隔
        
        # 第二步：按7天销量排序并评分（由高到低）
        valid_products.sort(key=lambda x: x['sales'], reverse=True)
        for i, product in enumerate(valid_products):
            product['sales_score'] = 100 - (i * score_step)  # 最高100分
        
        # 第三步：按达人数量排序并评分（由低到高）
        valid_products.sort(key=lambda x: x['influencers'])
        for i, product in enumerate(valid_products):
            product['influencer_score'] = 100 - (i * score_step)  # 最低数量得最高分
        
        # 第四步：计算加权总分
        results = []
        for product_data in valid_products:
            final_score = (product_data['sales_score'] * sales_weight + 
                          product_data['influencer_score'] * influencer_weight)
            
            results.append({
                'product_id': product_data['product'].get('product_id'),
                'product_name': product_data['product'].get('product_name'),
                'avg_price': product_data['product'].get('avg_price'),
                'total_sale_nd_cnt': product_data['product'].get('total_sale_nd_cnt'),
                'influencers_count': product_data['product'].get('influencers_count'),
                'product_rating': product_data['product'].get('product_rating'),
                'category': product_data['product'].get('category'),
                'sales_score': round(product_data['sales_score'], 2),
                'influencer_score': round(product_data['influencer_score'], 2),
                'final_score': round(final_score, 2),
                'cover_url': product_data['product'].get('cover_url'),  # 添加封面图URL
            })
        
        # 按总分排序
        results.sort(key=lambda x: x['final_score'], reverse=True)
        
        # 打印统计信息
        print(f"\n数据分析统计:")
        print(f"总商品数: {len(products)}")
        print(f"价格范围内的有效商品数: {total_products}")
        print(f"跳过商品数: {skipped_count}")
        print(f"\n评分权重:")
        print(f"销量得分权重: {sales_weight * 100}%")
        print(f"达人数量得分权重: {influencer_weight * 100}%")
        
        # 在返回结果之前，生成Excel报告
        self._generate_excel_report(results, price_range, sales_weight, influencer_weight)
        self._generate_txt_report(results)  # 添加生成TXT报告
        
        return results[:top_n]

    def _generate_excel_report(self, results: List[Dict], price_range: tuple, 
                              sales_weight: float, influencer_weight: float):
        """
        使用 openpyxl 生成Excel报告，确保图片正确嵌入单元格
        """
        try:
            from PIL import Image as PILImage  # 确保 Pillow 已安装
        except ImportError:
            print("错误: 请先安装 Pillow 库: pip install Pillow")
            return

        # 只取前50条数据
        top_50_results = results[:50]
        
        # 创建工作簿和工作表
        wb = Workbook()
        ws = wb.active
        ws.title = 'TOP 50分析结果'
        
        # 设置列标题
        headers = [
            '商品名称', '类别', '价格', '7天销量', '达人数量',
            '商品评分', '销量得分', '达人得分', '最终得分', '商品图片'
        ]
        
        # 写入标题行
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = cell.font.copy(bold=True)
            cell.fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置列宽
        ws.column_dimensions['A'].width = 30  # 商品名称
        ws.column_dimensions['B'].width = 12  # 类别
        ws.column_dimensions['C'].width = 10  # 价格
        ws.column_dimensions['D'].width = 12  # 7天销量
        ws.column_dimensions['E'].width = 12  # 达人数量
        for col in range(6, 10):
            ws.column_dimensions[get_column_letter(col)].width = 10  # 其他列
        ws.column_dimensions['J'].width = 25  # 图片列
        
        # 写入数据并插入图片
        for row_idx, product in enumerate(top_50_results, 2):
            # 设置行高
            ws.row_dimensions[row_idx].height = 150
            
            # 写入数据
            data = [
                product['product_name'],
                product['category'],
                product['avg_price'],
                product['total_sale_nd_cnt'],
                product['influencers_count'],
                product['product_rating'],
                f"{product['sales_score']:.2f}",
                f"{product['influencer_score']:.2f}",
                f"{product['final_score']:.2f}"
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # 下载并插入图片
            image_url = product['cover_url']
            if image_url:
                print(f"正在处理第 {row_idx-1}/50 张图片: {image_url}")
                try:
                    response = requests.get(image_url, timeout=10)
                    if response.status_code == 200:
                        image_bytes = BytesIO()
                        image_bytes.write(response.content)
                        
                        # 使用 PIL 处理图片
                        pil_image = PILImage.open(image_bytes)
                        
                        # 调整图片大小
                        target_width = 180
                        target_height = 200
                        pil_image.thumbnail((target_width, target_height), PILImage.LANCZOS)
                        
                        # 保存处理后的图片
                        output_bytes = BytesIO()
                        pil_image.save(output_bytes, format='PNG')
                        output_bytes.seek(0)
                        
                        # 创建 openpyxl 图片对象
                        img = Image(output_bytes)
                        
                        # 创建图片锚点
                        from_marker = AnchorMarker(
                            col=9,  # J列（图片列）
                            colOff=50000,
                            row=row_idx-1,
                            rowOff=50000
                        )
                        
                        to_marker = AnchorMarker(
                            col=10,
                            colOff=-50000,
                            row=row_idx,
                            rowOff=-50000
                        )
                        
                        img.anchor = TwoCellAnchor('twoCell', from_marker, to_marker)
                        ws.add_image(img)
                        print(f"✓ 图片 {row_idx-1} 处理成功")
                        
                    else:
                        print(f"× 下载图片失败: HTTP {response.status_code}")
                        
                except requests.exceptions.Timeout:
                    print(f"× 下载超时: {image_url}")
                except Exception as e:
                    print(f"× 处理图片时出错: {str(e)}")
                    import traceback
                    print(traceback.format_exc())
        
        # 冻结标题行
        ws.freeze_panes = 'A2'
        
        # 添加筛选条件说明
        info_row = len(top_50_results) + 3
        ws.cell(row=info_row, column=1, value='筛选条件:').font = Font(bold=True)
        ws.cell(row=info_row+1, column=1, value=f'价格范围: RM{price_range[0]}-{price_range[1]}')
        ws.cell(row=info_row+2, column=1, value=f'销量得分权重: {sales_weight * 100}%')
        ws.cell(row=info_row+3, column=1, value=f'达人得分权重: {influencer_weight * 100}%')
        
        # 保存文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = f'top50_analysis_{timestamp}.xlsx'
        wb.save(excel_file)
        
        print(f"\nTOP 50分析报告已生成: {excel_file}")

    def _generate_txt_report(self, results: List[Dict]):
        """
        生成TXT格式的简要报告，包含前50个商品的基本信息
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        txt_file = f'top50_brief_{timestamp}.txt'
        
        with open(txt_file, 'w', encoding='utf-8') as f:
            for i, product in enumerate(results[:50], 1):
                f.write(f"{i}. {product['product_name']}, "
                       f"{product['total_sale_nd_cnt']}, "
                       f"{product['influencers_count']}\n")
        
        print(f"\n简要报告已生成: {txt_file}")

    def _generate_excel_report_with_1688(self, results: List[Dict], 
                                       price_range: tuple,
                                       sales_weight: float,
                                       influencer_weight: float,
                                       method: int = 1):
        """
        生成包含1688匹配结果的Excel报告
        method: 1 表示方案一，2 表示方案二
        """
        # ... Excel初始化代码 ...

        # 初始化1688搜索
        self.alibaba_searcher.init_browser()

        filtered_results = []
        for product in results[:50]:
            shopee_image_url = product['cover_url']
            
            # 搜索1688商品
            alibaba_products = self.alibaba_searcher.search_by_image(shopee_image_url)
            
            if method == 1:
                # 方案一：保留所有商品，直接添加1688搜索结果
                product['alibaba_matches'] = alibaba_products[:3]  # 保存前3个结果
                filtered_results.append(product)
                
            else:
                # 方案二：检查图片相似度
                similar_products = []
                for alibaba_product in alibaba_products[:10]:
                    similarity = self.image_matcher.compare_images(
                        shopee_image_url, 
                        alibaba_product['image_url']
                    )
                    if similarity >= self.image_matcher.similarity_threshold:
                        similar_products.append({
                            **alibaba_product,
                            'similarity': similarity
                        })
                
                # 如果有两个以上高相似度的商品，保留这个商品
                if len(similar_products) >= 2:
                    product['alibaba_matches'] = similar_products
                    filtered_results.append(product)

        # 生成Excel报告
        wb = Workbook()
        ws = wb.active
        
        # 添加1688相关的列
        headers.extend(['1688商品1', '1688链接1', '1688商品2', '1688链接2', '1688商品3', '1688链接3'])
        
        # ... 写入数据 ...
        for row_idx, product in enumerate(filtered_results, 2):
            # ... 原有数据写入代码 ...
            
            # 写入1688匹配结果
            for i, match in enumerate(product.get('alibaba_matches', [])[:3], 1):
                img_col = len(headers) - 6 + (i-1)*2
                link_col = img_col + 1
                
                # 插入1688商品图片
                self._insert_image_to_cell(ws, row_idx, img_col, match['image_url'])
                
                # 插入1688商品链接
                ws.cell(row_idx, link_col, match['product_url'])

        # 保存文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = f'top50_analysis_with_1688_{timestamp}.xlsx'
        wb.save(excel_file)
        
        print(f"\nTOP 50分析报告（含1688匹配）已生成: {excel_file}")

def _insert_image_to_cell(ws, row, col, image_url):
    """在指定单元格中插入图片"""
    try:
        # 下载图片
        response = requests.get(image_url)
        img_data = BytesIO(response.content)
        
        # 创建PIL图像对象
        pil_img = PILImage.open(img_data)
        
        # 调整图片大小
        max_width, max_height = 120, 120
        pil_img.thumbnail((max_width, max_height), PILImage.LANCZOS)
        
        # 保存调整后的图片到内存
        img_byte_arr = BytesIO()
        pil_img.save(img_byte_arr, format=pil_img.format or 'PNG')
        img_byte_arr.seek(0)
        
        # 创建openpyxl图片对象
        img = Image(img_byte_arr)
        
        # 设置图片位置
        img.anchor = f'{get_column_letter(col)}{row}'
        
        # 调整单元格大小以适应图片
        ws.row_dimensions[row].height = 90  # 设置行高
        ws.column_dimensions[get_column_letter(col)].width = 18  # 设置列宽
        
        # 添加图片到工作表
        ws.add_image(img)
        
    except Exception as e:
        print(f"插入图片时出错: {str(e)}")
        ws.cell(row, col, "图片加载失败")

def test_single_product():
    """测试单个商品的1688识图功能"""
    test_product = {
        "product_id": "1730761319617628044",
        "product_name": "[4 pek TISU PANDA POPI] Popi Tisu Gantung...",
        "cover_url": "https://cdn.echotik.shop/b687bd5049700729ced9c36c73f86e14/67d3cb91/product-cover/175/1729527232466552025_0.jpeg",
        "category": "家居用品",
        "avg_price": "RM10.00",
        "total_sale_nd_cnt": "39.4K",
        "influencers_count": "349",
        "product_rating": "4.8"
    }
    
    print("\n=== 开始测试两种方案 ===")
    print(f"测试商品: {test_product['product_name']}")
    print(f"商品图片: {test_product['cover_url']}")
    
    # 初始化匹配器和搜索器
    matcher = ImageMatcher()
    
    try:
        # 初始化浏览器
        matcher.init_browser()
        
        # 执行图片搜索
        print("\n1. 搜索1688商品...")
        results = matcher.search_1688_by_image(test_product['cover_url'])
        
        # 方案一：直接保存前3个结果
        print("\n=== 方案一结果 ===")
        print("保存前3个搜索结果:")
        scheme1_results = results[:3]
        for i, product in enumerate(scheme1_results, 1):
            print(f"\n{i}. 商品标题: {product['title']}")
            print(f"   商品价格: {product['price']}")
            print(f"   商品链接: {product['product_url']}")
            print(f"   商品图片: {product['image_url']}")
        
        # 方案二：计算相似度并筛选
        print("\n=== 方案二结果 ===")
        print("计算前10个结果的相似度:")
        similar_products = []
        
        for i, product in enumerate(results[:10], 1):
            print(f"\n处理第 {i} 个商品:")
            print(f"标题: {product['title']}")
            
            # 计算相似度
            similarity = matcher.compare_images(
                test_product['cover_url'],
                product['image_url']
            )
            
            print(f"相似度: {similarity:.4f}")
            
            # 如果相似度超过阈值，保存结果
            if similarity >= matcher.similarity_threshold:
                similar_products.append({
                    **product,
                    'similarity': similarity
                })
                print("✓ 相似度达标，已保存")
            else:
                print("× 相似度未达标，跳过")
        
        # 判断是否为有效商品（两个以上高相似度）
        is_valid = len(similar_products) >= 2
        print(f"\n商品有效性: {'✓ 有效' if is_valid else '× 无效'}")
        
        if is_valid:
            print(f"找到 {len(similar_products)} 个相似商品:")
            # 按相似度排序
            similar_products.sort(key=lambda x: x['similarity'], reverse=True)
            for i, product in enumerate(similar_products, 1):
                print(f"\n{i}. 相似度: {product['similarity']:.4f}")
                print(f"   标题: {product['title']}")
                print(f"   价格: {product['price']}")
                print(f"   链接: {product['product_url']}")
                print(f"   图片: {product['image_url']}")
        
        # 生成测试报告
        print("\n=== 生成Excel报告 ===")
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "方案一结果"
        ws2 = wb.create_sheet("方案二结果")
        
        # 设置列宽
        for ws in [ws1, ws2]:
            for col in range(1, 20):
                ws.column_dimensions[get_column_letter(col)].width = 15
        
        # 定义表头
        headers = [
            '商品ID', '商品名称', '类别', '平均价格', '总销量',
            '达人数量', '商品评分', '商品图片',
            '1688商品1', '1688链接1',
            '1688商品2', '1688链接2',
            '1688商品3', '1688链接3'
        ]
        
        # 写入表头
        for ws in [ws1, ws2]:
            for col, header in enumerate(headers, 1):
                ws.cell(1, col, header)
                ws.cell(1, col).font = Font(bold=True)
        
        # 写入测试数据
        row = 2
        ws1.cell(row, 1, test_product['product_id'])
        ws1.cell(row, 2, test_product['product_name'])
        ws1.cell(row, 3, test_product['category'])
        ws1.cell(row, 4, test_product['avg_price'])
        ws1.cell(row, 5, test_product['total_sale_nd_cnt'])
        ws1.cell(row, 6, test_product['influencers_count'])
        ws1.cell(row, 7, test_product['product_rating'])
        
        # 插入商品原图
        _insert_image_to_cell(ws1, row, 8, test_product['cover_url'])
        
        # 方案一：插入前3个1688商品的图片和链接
        for i, product in enumerate(scheme1_results, 1):
            img_col = 8 + i*2 - 1  # 图片列
            link_col = img_col + 1  # 链接列
            
            # 插入1688商品图片
            _insert_image_to_cell(ws1, row, img_col, product['image_url'])
            
            # 插入1688商品链接
            ws1.cell(row, link_col, product['product_url'])
        
        # 方案二：如果有效，创建新工作表
        if is_valid:
            ws2 = wb.create_sheet("方案二结果")
            
            # 复制相同的表头和格式
            for col, header in enumerate(headers, 1):
                ws2.cell(1, col, header)
                ws2.cell(1, col).font = Font(bold=True)
                ws2.column_dimensions[get_column_letter(col)].width = 15
            
            # 写入相同的测试数据
            row = 2
            ws2.cell(row, 1, test_product['product_id'])
            ws2.cell(row, 2, test_product['product_name'])
            ws2.cell(row, 3, test_product['category'])
            ws2.cell(row, 4, test_product['avg_price'])
            ws2.cell(row, 5, test_product['total_sale_nd_cnt'])
            ws2.cell(row, 6, test_product['influencers_count'])
            ws2.cell(row, 7, test_product['product_rating'])
            
            # 插入商品原图
            _insert_image_to_cell(ws2, row, 8, test_product['cover_url'])
            
            # 插入高相似度的1688商品
            for i, product in enumerate(similar_products[:3], 1):
                img_col = 8 + i*2 - 1
                link_col = img_col + 1
                
                # 插入1688商品图片
                _insert_image_to_cell(ws2, row, img_col, product['image_url'])
                
                # 插入1688商品链接和相似度信息
                ws2.cell(row, link_col, 
                    f"相似度: {product['similarity']:.4f}\n"
                    f"链接: {product['product_url']}")
        
        # 保存Excel文件
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_file = f'test_results_{timestamp}.xlsx'
        wb.save(excel_file)
        print(f"\nExcel报告已生成: {excel_file}")
        
    except Exception as e:
        print(f"\n测试过程中出现错误: {str(e)}")
        import traceback
        print(traceback.format_exc())
    finally:
        # 确保浏览器被关闭
        matcher.close_browser()

def main():
    try:
        # 1. 初始化数据分析器
        analyzer = DataAnalyzer()
        
        # 2. 分析并筛选TOP50商品
        print("\n1. 开始分析商品数据...")
        top50_results = analyzer.analyze_products(
            price_range=(10, 1000),
            sales_weight=0.6,
            influencer_weight=0.4,
            top_n=50
        )
        
        if not top50_results:
            print("未找到符合条件的商品")
            return
            
        print(f"✓ 找到 {len(top50_results)} 个符合条件的商品")
        
        # 3. 初始化匹配器并开始1688匹配
        print(f"\n2. 开始1688商品匹配...")
        matcher = ImageMatcher()
        
        try:
            matcher.init_browser()
            
            # 4. 创建Excel工作簿
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "方案一结果"
            ws2 = wb.create_sheet("方案二结果")
            
            # 5. 处理每个TOP50商品
            valid_products = []  # 存储方案二的有效商品
            processed_count = 0  # 记录成功处理的商品数
            
            for row, product in enumerate(top50_results, 2):
                try:
                    print(f"\n{'='*50}")
                    print(f"处理第 {row-1}/{len(top50_results)} 个商品: {product['product_name']}")
                    print(f"商品URL: {product['cover_url']}")
                    
                    # 搜索1688商品
                    try:
                        results = matcher.search_1688_by_image(product['cover_url'])
                        if not results:
                            print(f"× 未找到匹配商品，跳过")
                            continue
                            
                        print(f"✓ 找到 {len(results)} 个匹配商品")
                        processed_count += 1
                        
                        # 写入基础数据（两个表都写入）
                        for ws in [ws1, ws2]:
                            ws.cell(row, 1, product['product_id'])
                            ws.cell(row, 2, product['product_name'])
                            ws.cell(row, 3, product.get('category', ''))
                            ws.cell(row, 4, product.get('avg_price', ''))
                            ws.cell(row, 5, product.get('total_sale_nd_cnt', ''))
                            ws.cell(row, 6, product.get('influencers_count', ''))
                            ws.cell(row, 7, product.get('product_rating', ''))
                            
                            # 插入商品原图
                            _insert_image_to_cell(ws, row, 8, product['cover_url'])
                        
                        # 方案一：写入前3个结果
                        scheme1_results = results[:3]
                        for i, result in enumerate(scheme1_results, 1):
                            img_col = 8 + i*2 - 1
                            link_col = img_col + 1
                            _insert_image_to_cell(ws1, row, img_col, result['image_url'])
                            ws1.cell(row, link_col, result['product_url'])
                        
                        # 方案二：计算相似度
                        similar_products = []
                        for result in results[:10]:
                            similarity = matcher.compare_images(
                                product['cover_url'],
                                result['image_url']
                            )
                            if similarity >= matcher.similarity_threshold:
                                similar_products.append({**result, 'similarity': similarity})
                        
                        # 如果有两个以上高相似度商品，写入方案二表格
                        if len(similar_products) >= 2:
                            similar_products.sort(key=lambda x: x['similarity'], reverse=True)
                            for i, result in enumerate(similar_products[:3], 1):
                                img_col = 8 + i*2 - 1
                                link_col = img_col + 1
                                _insert_image_to_cell(ws2, row, img_col, result['image_url'])
                                ws2.cell(row, link_col, 
                                    f"相似度: {result['similarity']:.4f}\n"
                                    f"链接: {result['product_url']}")
                            valid_products.append(product)
                        
                        # 每处理完一个商品，打印进度
                        print(f"\n当前进度: {row-1}/{len(top50_results)} 完成")
                        
                        # 每处理完一个商品暂停一下，避免请求过快
                        time.sleep(2)
                        
                    except Exception as e:
                        print(f"× 搜索商品时出错: {str(e)}")
                        print("跳过当前商品，继续处理下一个")
                        continue
                    
                except Exception as e:
                    print(f"× 处理商品时出错: {str(e)}")
                    print("继续处理下一个商品")
                    continue
            
            # 6. 保存最终Excel文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = f'top50_analysis_{timestamp}.xlsx'
            wb.save(excel_file)
            
            print(f"\n分析完成！")
            print(f"共处理 {processed_count}/{len(top50_results)} 个商品")
            print(f"方案一：保存了所有处理成功的商品匹配结果")
            print(f"方案二：找到 {len(valid_products)} 个有效商品")
            print(f"Excel报告已生成: {excel_file}")
            
            return excel_file
            
        finally:
            matcher.close_browser()
            
    except Exception as e:
        print(f"\n处理过程中出现错误: {str(e)}")
        import traceback
        print(traceback.format_exc())

def analyze_product_data(data_file: str, strategy: str = 'all') -> tuple:
    """分析商品数据并生成报告"""
    try:
        print(f"\n=== 开始分析商品数据 ===")
        print(f"数据文件: {data_file}")
        print(f"分析策略: {strategy}")
        
        # 1. 初始化数据分析器
        data_dir = os.path.dirname(data_file) if os.path.dirname(data_file) else 'data'
        analyzer = DataAnalyzer(data_dir=data_dir)
        
        # 2. 分析并筛选TOP50商品
        print("\n1. 开始分析商品数据...")
        top50_results = analyzer.analyze_products(
            price_range=(10, 1000),
            sales_weight=0.6,
            influencer_weight=0.4,
            top_n=50
        )
        
        if not top50_results:
            raise Exception("未找到符合条件的商品")
            
        print(f"✓ 找到 {len(top50_results)} 个符合条件的商品")
        
        # 测试模式：限制处理5条数据
        test_mode = False  # 设置为 False 可以处理全部数据
        if test_mode:
            print("\n⚠️ 测试模式：只处理前5条数据")
            top50_results = top50_results[:5]
        
        # 3. 初始化匹配器并开始1688匹配
        print(f"\n2. 开始1688商品匹配...")
        matcher = ImageMatcher()
        
        # 创建结果JSON数据结构
        analysis_results = {
            "products": [],
            "summary": {
                "total_products": len(top50_results),
                "processed_count": 0,
                "scheme1_count": 0,
                "scheme2_count": 0
            }
        }
        
        try:
            matcher.init_browser()
            
            # 4. 创建Excel工作簿
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "方案一结果"
            ws2 = wb.create_sheet("方案二结果")
            
            # 5. 处理每个商品
            valid_products = []
            processed_count = 0
            
            for row, product in enumerate(top50_results, 2):
                try:
                    print(f"\n{'='*50}")
                    print(f"处理第 {row-1}/{len(top50_results)} 个商品: {product['product_name']}")
                    print(f"商品URL: {product['cover_url']}")
                    
                    # 创建当前商品的结果对象
                    product_result = {
                        "product_id": product['product_id'],
                        "product_name": product['product_name'],
                        "cover_url": product['cover_url'],
                        "category": product.get('category', ''),
                        "avg_price": product.get('avg_price', ''),
                        "total_sale_nd_cnt": product.get('total_sale_nd_cnt', ''),
                        "influencers_count": product.get('influencers_count', ''),
                        "product_rating": product.get('product_rating', ''),
                        "matches_1": [],  # 方案一结果
                        "matches_2": []   # 方案二结果
                    }
                    
                    # 搜索1688商品
                    try:
                        results = matcher.search_1688_by_image(product['cover_url'])
                        if not results:
                            print(f"× 未找到匹配商品，跳过")
                            continue
                            
                        print(f"✓ 找到 {len(results)} 个匹配商品")
                        processed_count += 1
                        analysis_results["summary"]["processed_count"] += 1
                        
                        # 写入基础数据（两个表都写入）
                        for ws in [ws1, ws2]:
                            ws.cell(row, 1, product['product_id'])
                            ws.cell(row, 2, product['product_name'])
                            ws.cell(row, 3, product.get('category', ''))
                            ws.cell(row, 4, product.get('avg_price', ''))
                            ws.cell(row, 5, product.get('total_sale_nd_cnt', ''))
                            ws.cell(row, 6, product.get('influencers_count', ''))
                            ws.cell(row, 7, product.get('product_rating', ''))
                            
                            # 插入商品原图
                            _insert_image_to_cell(ws, row, 8, product['cover_url'])
                        
                        # 方案一：写入前3个结果
                        scheme1_results = results[:3]
                        for i, result in enumerate(scheme1_results, 1):
                            img_col = 8 + i*2 - 1
                            link_col = img_col + 1
                            _insert_image_to_cell(ws1, row, img_col, result['image_url'])
                            ws1.cell(row, link_col, result['product_url'])
                            
                            # 保存到JSON结果
                            product_result["matches_1"].append({
                                "image_url": result['image_url'],
                                "product_url": result['product_url'],
                                "title": result.get('title', ''),
                                "price": result.get('price', '')
                            })
                        
                        # 方案二：计算相似度
                        similar_products = []
                        for result in results[:10]:
                            similarity = matcher.compare_images(
                                product['cover_url'],
                                result['image_url']
                            )
                            if similarity >= matcher.similarity_threshold:
                                similar_products.append({**result, 'similarity': similarity})
                                
                                # 保存到JSON结果
                                product_result["matches_2"].append({
                                    "image_url": result['image_url'],
                                    "product_url": result['product_url'],
                                    "title": result.get('title', ''),
                                    "price": result.get('price', ''),
                                    "similarity": float(similarity)
                                })
                        
                        # 如果有两个以上高相似度商品，写入方案二表格
                        if len(similar_products) >= 2:
                            similar_products.sort(key=lambda x: x['similarity'], reverse=True)
                            for i, result in enumerate(similar_products[:3], 1):
                                img_col = 8 + i*2 - 1
                                link_col = img_col + 1
                                _insert_image_to_cell(ws2, row, img_col, result['image_url'])
                                ws2.cell(row, link_col, 
                                    f"相似度: {result['similarity']:.4f}\n"
                                    f"链接: {result['product_url']}")
                            valid_products.append(product)
                            analysis_results["summary"]["scheme2_count"] += 1
                        
                        # 添加到分析结果
                        if product_result["matches_1"] or product_result["matches_2"]:
                            analysis_results["products"].append(product_result)
                            if product_result["matches_1"]:
                                analysis_results["summary"]["scheme1_count"] += 1
                        
                    except Exception as e:
                        print(f"× 搜索商品时出错: {str(e)}")
                        print("跳过当前商品，继续处理下一个")
                        continue
                    
                except Exception as e:
                    print(f"× 处理商品时出错: {str(e)}")
                    continue
            
            # 6. 保存Excel文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_file = f'top50_analysis_{timestamp}.xlsx'
            wb.save(excel_file)
            
            # 7. 保存JSON结果文件到analysis目录
            analysis_dir = 'analysis'
            os.makedirs(analysis_dir, exist_ok=True)
            json_file = os.path.join(analysis_dir, f'analysis_results_{timestamp}.json')
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(analysis_results, f, ensure_ascii=False, indent=2)
            
            print(f"\n分析完成！")
            print(f"共处理 {processed_count}/{len(top50_results)} 个商品")
            print(f"方案一：保存了所有处理成功的商品匹配结果")
            print(f"方案二：找到 {len(valid_products)} 个有效商品")
            print(f"Excel报告已生成: {excel_file}")
            print(f"JSON结果已生成: {json_file}")
            
            return excel_file, json_file
            
        finally:
            if matcher:
                matcher.close_browser()
            
    except Exception as e:
        print(f"\n分析过程出错: {str(e)}")
        raise

def analyze_1688(excel_file: str) -> None:
    """对已生成的Excel文件进行1688商品匹配分析"""
    # 这个函数可以根据需要实现，或者直接返回
    pass

if __name__ == "__main__":
    main()
    # test_single_product()  # 取消注释此行可以运行测试 